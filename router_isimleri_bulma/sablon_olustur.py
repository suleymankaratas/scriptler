"""IP listesi icin Excel sablonu olusturur."""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Router Listesi"

headers = ["IP Adresi", "Kullanici Adi", "Sifre", "Port (bos=23)", "Router Adi", "Durum"]
ws.append(headers)

header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
for col_idx, _ in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Ornek satir (silinebilir)
ws.append(["192.168.1.1", "admin", "sifre123", "", "", ""])

widths = [16, 16, 16, 14, 30, 40]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"

out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "router_listesi.xlsx")
wb.save(out_path)
print("Kaydedildi:", out_path)
