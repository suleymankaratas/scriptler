"""
router_listesi.xlsx dosyasindaki IP'lere Telnet ile baglanip
router isimlerini (prompt/hostname) bulup ayni dosyaya yazar.

Kullanim:
    python router_isimleri_bul.py

Excel dosyasi sutunlari:
    IP Adresi | Kullanici Adi | Sifre | Port (bos=23) | Router Adi | Durum

Not: Python 3.13+ 'telnetlib' modulunu kaldirdigi icin burada
sadece standart kutuphaneyle (socket) minimal bir telnet istemcisi
yazildi. Ek kutuphane kurmaya gerek yok (openpyxl haric).
"""

import os
import socket
import sys
import time
import re
import openpyxl

# Script'in bulundugu klasordeki Excel dosyasini kullanir; klasor
# tasinsa/kopyalansa bile calismaya devam eder (sabit yol degil).
EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "router_listesi.xlsx")

# --- Telnet IAC (negotiation) sabitleri ---
IAC, WILL, WONT, DO, DONT, SB, SE = 255, 251, 252, 253, 254, 250, 240


class MiniTelnet:
    """Sadece bu script icin yazilmis, bagimliliksiz minimal telnet istemcisi."""

    def __init__(self, host, port, timeout=8):
        self.sock = socket.create_connection((host, port), timeout=timeout)
        self.sock.settimeout(timeout)
        self._raw_buffer = b""
        self.text_buffer = ""

    def _process(self, chunk):
        data = self._raw_buffer + chunk
        out = bytearray()
        i, n = 0, len(data)
        while i < n:
            b = data[i]
            if b == IAC:
                if i + 1 >= n:
                    break
                cmd = data[i + 1]
                if cmd in (WILL, WONT, DO, DONT):
                    if i + 2 >= n:
                        break
                    opt = data[i + 2]
                    try:
                        if cmd == DO:
                            self.sock.sendall(bytes([IAC, WONT, opt]))
                        elif cmd == WILL:
                            self.sock.sendall(bytes([IAC, DONT, opt]))
                    except OSError:
                        pass
                    i += 3
                    continue
                elif cmd == SB:
                    end = data.find(bytes([IAC, SE]), i)
                    if end == -1:
                        break
                    i = end + 2
                    continue
                elif cmd == IAC:
                    out.append(IAC)
                    i += 2
                    continue
                else:
                    i += 2
                    continue
            else:
                out.append(b)
                i += 1
        self._raw_buffer = data[i:]
        self.text_buffer += out.decode("utf-8", errors="replace")

    def read_until(self, patterns, timeout=8):
        end_time = time.time() + timeout
        while True:
            low = self.text_buffer.lower()
            for p in patterns:
                if p.lower() in low:
                    return self.text_buffer
            remaining = end_time - time.time()
            if remaining <= 0:
                return self.text_buffer
            self.sock.settimeout(min(remaining, 1.0))
            try:
                chunk = self.sock.recv(4096)
            except socket.timeout:
                continue
            except OSError:
                return self.text_buffer
            if not chunk:
                return self.text_buffer
            self._process(chunk)

    def write(self, s):
        self.sock.sendall(s.encode("utf-8", errors="replace") + b"\r\n")

    def close(self):
        try:
            self.sock.close()
        except OSError:
            pass


# Kapsanan prompt formatlari:
#   Cisco/MikroTik tarzi:  RouterName>   RouterName#
#   Huawei/H3C tarzi:      <RouterName>  [RouterName]
PROMPT_RE = re.compile(r"^[<\[]?([A-Za-z0-9_\-\.@]{1,80})[>\]#]\s*$")


def _find_prompt(text):
    for line in reversed(text.splitlines()):
        m = PROMPT_RE.match(line.strip("\r").strip())
        if m:
            return m.group(1)
    return None


def get_router_name(ip, username, password, port=23, timeout=8):
    tn = None
    try:
        tn = MiniTelnet(ip, port, timeout=timeout)

        banner = tn.read_until(["username:", "login:", "password:", ">", "#"], timeout=timeout)

        if "username" in banner.lower() or "login" in banner.lower():
            tn.write(username)
            banner = tn.read_until(["password:", ">", "#"], timeout=timeout)

        if "password" in banner.lower():
            tn.write(password)
            banner = tn.read_until(
                [">", "#", "denied", "incorrect", "failed", "invalid"], timeout=timeout
            )

        low = banner.lower()
        if any(k in low for k in ["denied", "incorrect", "failed", "invalid"]):
            return None, "Giris reddedildi (kullanici adi/sifre hatali olabilir)"

        name = _find_prompt(banner)
        if not name:
            tn.write("")
            more = tn.read_until([">", "#"], timeout=3)
            name = _find_prompt(more)

        if name:
            return name, "OK"
        return None, "Router adi (prompt) tespit edilemedi"

    except socket.timeout:
        return None, "Zaman asimi (baglanti kurulamadi)"
    except ConnectionRefusedError:
        return None, "Baglanti reddedildi (telnet portu kapali olabilir)"
    except OSError as e:
        return None, f"Baglanti hatasi: {e}"
    except Exception as e:  # noqa: BLE001
        return None, f"Hata: {e}"
    finally:
        if tn:
            tn.close()


def main():
    # Test icin: "python router_isimleri_bul.py 15" -> sadece ilk 15 IP'yi isler
    # Basarisizlari tekrar denemek icin: "python router_isimleri_bul.py --retry-failed"
    retry_failed = "--retry-failed" in sys.argv
    numeric_args = [a for a in sys.argv[1:] if a != "--retry-failed"]
    limit = int(numeric_args[0]) if numeric_args else None

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    col = {name: idx + 1 for idx, name in enumerate(headers) if name}

    required = ["IP Adresi", "Kullanici Adi", "Sifre", "Router Adi", "Durum"]
    for r in required:
        if r not in col:
            raise SystemExit(f"Excel'de '{r}' sutunu bulunamadi. Sablonu bozmadigindan emin ol.")

    processed = 0
    for row in range(2, ws.max_row + 1):
        if limit is not None and processed >= limit:
            print(f"\nTest limiti ({limit}) doldu, durduruluyor.")
            break

        ip_cell = ws.cell(row=row, column=col["IP Adresi"]).value
        if not ip_cell or not str(ip_cell).strip():
            continue

        existing_status = ws.cell(row=row, column=col["Durum"]).value
        if existing_status == "OK":
            continue  # daha once basariyla islendi, atla
        if existing_status and not retry_failed:
            continue  # daha once denenmis (basarisiz) ve --retry-failed verilmemis, atla

        processed += 1
        ip = str(ip_cell).strip()
        username = str(ws.cell(row=row, column=col["Kullanici Adi"]).value or "")
        password = str(ws.cell(row=row, column=col["Sifre"]).value or "")
        port_val = ws.cell(row=row, column=col.get("Port (bos=23)", 0)).value if col.get("Port (bos=23)") else None
        port = int(port_val) if port_val else 23

        print(f"[Satir {row}] {ip}:{port} deneniyor...")
        name, status = get_router_name(ip, username, password, port=port)
        ws.cell(row=row, column=col["Router Adi"]).value = name or ""
        ws.cell(row=row, column=col["Durum"]).value = status
        print(f"    -> Ad: {name}   Durum: {status}")

        wb.save(EXCEL_PATH)  # her satirdan sonra kaydet ki kesinti olursa veri kaybolmasin

    print("\nTamamlandi. Dosya guncellendi:", EXCEL_PATH)


if __name__ == "__main__":
    main()
