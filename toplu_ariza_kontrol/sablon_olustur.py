# -*- coding: utf-8 -*-
"""IP listesi icin Excel sablonu olusturur."""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Ariza Kontrol"

headers = [
    "Sorun / Genel Durum",
    "IP Adresi",
    "Okul / Aciklama",
    "Kullanici Adi",
    "Sifre",
    "Port (bos=23)",
    "Erisim Durumu",
    "Marka",
    "VLAN1 Internet (Yeni Router)",
    "VLAN10 Internet (Yonetim/AP)",
    "VLAN11 Internet (Tahta)",
    "VLAN20 Internet (Idare)",
    "VLAN30 Internet (Ogretmen)",
    "VLAN40 Internet (BT Sinifi)",
    "VLAN50 Internet (Ogrenci Tablet)",
    "VLAN60 Internet (Aktivasyon)",
    "ARP Cihaz Sayisi",
    "WLC Erisimi",
    "EBA Erisimi",
    "VLAN1 Option43 (Kablosuz)",
    "VLAN10 Option43 (Kablosuz)",
    "Reel IP Internet Cikisi",
    "Harici ADSL Modem",
    "NAT Calisiyor mu",
    "Interface Durumu (DOWN olanlar)",
    "Donanim/Log Alarmi (SRU-BPDU vb.)",
    "Not",
]
ws.append(headers)

header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
for col_idx, _ in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

# Ornek satir (silinebilir) - IP alani kendi listenle degistirilecek
ws.append(["", "192.168.1.1", "Ornek Okul", "suleyman.karatas@default", "Skrts123"] + [""] * 21)

widths = [40, 16, 26, 22, 14, 12, 16, 10, 14, 16, 16, 16, 16, 16, 18, 16, 14, 14, 14, 18, 18, 20, 16, 14, 22, 18, 30]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 30

out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ariza_listesi.xlsx")
wb.save(out_path)
print("Kaydedildi:", out_path)
