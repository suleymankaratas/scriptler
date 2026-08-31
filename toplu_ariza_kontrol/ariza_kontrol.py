# -*- coding: utf-8 -*-
"""
ariza_listesi.xlsx dosyasindaki IP'lere Telnet ile baglanip "Ariza Kontrol
Asamalari.docx" belgesindeki adimlari sirayla uygular:

  1) Telnet ile giris yapilabiliyor mu (erisim var mi, yoksa hangi hata)
  2) "display/show ip interface brief" ile VLAN interface IP'leri okunur
     (cihaz Huawei mi Cisco mu oldugu buradan otomatik anlasilir)
  3) VLAN 10/11/20/30/40/50/60 interface IP'lerinden 4.2.2.2'ye ping atilir
     (her VLAN'in interneti ayri ayri raporlanir)
  4) ARP tablosu okunur, kac cihaz bagli oldugu sayilir
  5) VLAN11 (Etkilesimli Tahta) uzerinden EBA/WLC sunucusuna (10.201.129.233)
     erisim kontrol edilir
  6) Genel not: ham "display current-configuration" satir sayisi gibi kaba
     bir bilgi + hata/aciklama notlari

Kullanim:
    python ariza_kontrol.py                  -> tum satirlari isler
    python ariza_kontrol.py 15                -> sadece ilk 15 satiri isler (test)
    python ariza_kontrol.py --retry-failed    -> erisim saglanamayan satirlari tekrar dener
"""

import os
import re
import socket
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
import openpyxl
from openpyxl.styles import Font, PatternFill

EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ariza_listesi.xlsx")
PING_HEDEFI = "4.2.2.2"
# Belgedeki iki ayri kontrol, iki farkli sunucuya ve iki farkli kaynak
# VLAN'a gidiyor:
#   WLC: "AP'ler WLC'ye gidiyor mu?"          -> VLAN10 (Yonetim/AP) IP'sinden
#   EBA: "ET'ler EBA sunucularina gidiyor mu?" -> VLAN11 (Tahta) IP'sinden
WLC_IP = "10.201.129.235"
EBA_IP = "10.201.129.233"

# Tum routerlarda ayni giris bilgisi kullanildigi icin varsayilan olarak
# tanimlandi. Excel'de "Kullanici Adi"/"Sifre" hucresi bos birakilan
# satirlarda bu bilgiler kullanilir; farkli bir cihaz icin Excel'e o
# satira ozel kullanici adi/sifre girmek yeterlidir (Excel'deki deger
# varsa o oncelikli olur).
VARSAYILAN_KULLANICI_ADI = "suleyman.karatas@default"
VARSAYILAN_SIFRE = "Skrts123"

# VLAN no -> (aciklama, excel sutun adi)
# NOT: VLAN1 sadece "yeni" Huawei routerlarda (Vlanif1 seklinde) bulunur;
# "eski" Huawei/Cisco routerlarda bu VLAN tanimli degildir. Bu satirlarda
# ilgili hucreye otomatik "-" yazilir ve sorun olarak sayilmaz.
VLAN_TANIMLARI = {
    1: "VLAN1 Internet (Yeni Router)",
    10: "VLAN10 Internet (Yonetim/AP)",
    11: "VLAN11 Internet (Tahta)",
    20: "VLAN20 Internet (Idare)",
    30: "VLAN30 Internet (Ogretmen)",
    40: "VLAN40 Internet (BT Sinifi)",
    50: "VLAN50 Internet (Ogrenci Tablet)",
    60: "VLAN60 Internet (Aktivasyon)",
}

VLAN_KISA_ISIM = {
    1: "VLAN1-YeniRouter",
    10: "VLAN10-Yonetim",
    11: "VLAN11-Tahta",
    20: "VLAN20-Idare",
    30: "VLAN30-Ogretmen",
    40: "VLAN40-BT",
    50: "VLAN50-Tablet",
    60: "VLAN60-Aktivasyon",
}

# --- Telnet IAC (negotiation) sabitleri ---
IAC, WILL, WONT, DO, DONT, SB, SE = 255, 251, 252, 253, 254, 250, 240


class MiniTelnet:
    """Bagimliliksiz minimal telnet istemcisi (Python 3.13+ telnetlib'i kaldirdigi icin)."""

    def __init__(self, host, port, timeout=8):
        self.sock = socket.create_connection((host, port), timeout=timeout)
        self.sock.settimeout(timeout)
        self._raw_buffer = b""
        self.text_buffer = ""

    def _process(self, chunk):
        data = self._raw_buffer + chunk
        out = bytearray()
        i, n = 0, len(data)
        while i < n:
            b = data[i]
            if b == IAC:
                if i + 1 >= n:
                    break
                cmd = data[i + 1]
                if cmd in (WILL, WONT, DO, DONT):
                    if i + 2 >= n:
                        break
                    opt = data[i + 2]
                    try:
                        if cmd == DO:
                            self.sock.sendall(bytes([IAC, WONT, opt]))
                        elif cmd == WILL:
                            self.sock.sendall(bytes([IAC, DONT, opt]))
                    except OSError:
                        pass
                    i += 3
                    continue
                elif cmd == SB:
                    end = data.find(bytes([IAC, SE]), i)
                    if end == -1:
                        break
                    i = end + 2
                    continue
                elif cmd == IAC:
                    out.append(IAC)
                    i += 2
                    continue
                else:
                    i += 2
                    continue
            else:
                out.append(b)
                i += 1
        self._raw_buffer = data[i:]
        self.text_buffer += out.decode("utf-8", errors="replace")

    def read_until(self, patterns, timeout=8):
        end_time = time.time() + timeout
        while True:
            low = self.text_buffer.lower()
            for p in patterns:
                if p.lower() in low:
                    return self.text_buffer
            remaining = end_time - time.time()
            if remaining <= 0:
                return self.text_buffer
            self.sock.settimeout(min(remaining, 1.0))
            try:
                chunk = self.sock.recv(4096)
            except socket.timeout:
                continue
            except OSError:
                return self.text_buffer
            if not chunk:
                return self.text_buffer
            self._process(chunk)

    def drain(self):
        """Sokette bekleyen (onceki komuttan kalma, gec gelmis) veriyi okuyup
        atar. Bu yapilmazsa bir onceki komutun gecikmis yaniti, sonraki
        komutun yanitiyla karisip yanlis sonuc/erken donme sebebi olabilir."""
        self.sock.settimeout(0.4)
        try:
            while True:
                chunk = self.sock.recv(4096)
                if not chunk:
                    break
                self._process(chunk)
        except socket.timeout:
            pass
        except OSError:
            pass
        self.text_buffer = ""
        self._raw_buffer = b""

    def write(self, s):
        self.sock.sendall(s.encode("utf-8", errors="replace") + b"\r\n")

    def close(self):
        try:
            self.sock.close()
        except OSError:
            pass


PACKET_LOSS_RE = re.compile(r"(\d+(?:\.\d+)?)\s*%\s*packet loss")
# Iki farkli interface adlandirma bicimini kapsar:
#   eski Huawei / Cisco sub-interface:  GigabitEthernet0/0/1.10   10.6.148.129/26 ...
#   yeni Huawei VLAN interface:         Vlanif10                  10.6.148.129/26 ...
IFACE_IP_RE = re.compile(
    r"(?:vlanif(\d{1,4})|\.(\d{1,3}))\s+(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})(?:/\d+)?\s",
    re.IGNORECASE,
)
IP_RE = re.compile(r"\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}")

# Router prompt formatlari:
#   Cisco/MikroTik tarzi:  RouterName>   RouterName#
#   Huawei/H3C tarzi:      <RouterName>  [RouterName]
PROMPT_RE = re.compile(r"^[<\[]?([A-Za-z0-9_\-\.@]{1,80})[>\]#]\s*$")


def _find_prompt(text):
    """Banner/komut ciktisindaki son satirdan cihazin prompt'unu (genelde
    okul/kurum adini icerir) cikarir."""
    for line in reversed(text.splitlines()):
        m = PROMPT_RE.match(line.strip("\r").strip())
        if m:
            return m.group(1)
    return None


def _login(tn, username, password, timeout):
    """Telnet oturumuna giris yapar. (True, banner) / (False, hata_aciklamasi) doner."""
    banner = tn.read_until(["username:", "login:", "password:", ">", "#"], timeout=timeout)

    if "username" in banner.lower() or "login" in banner.lower():
        tn.write(username)
        banner = tn.read_until(["password:", ">", "#"], timeout=timeout)

    if "password" in banner.lower():
        tn.write(password)
        banner = tn.read_until(
            [">", "#", "denied", "incorrect", "failed", "invalid"], timeout=timeout
        )

    low = banner.lower()
    if any(k in low for k in ["denied", "incorrect", "failed", "invalid"]):
        return False, "Kullanici adi/sifre reddedildi"
    if not (">" in banner or "#" in banner):
        return False, "Giris sonrasi komut istemi (prompt) alinamadi"
    return True, banner


def _run_cmd(tn, command, timeout=10):
    # onceki komutun gec gelmis kalintisini (sokette bekleyen veri dahil)
    # temizle; aksi halde eski bir ">"/"#" yuzunden yeni komutun yaniti
    # gelmeden fonksiyon erken donebilir.
    tn.drain()
    tn.write(command)
    return tn.read_until([">", "#"], timeout=timeout)


def _run_config_cmd(tn, command, expected_hostname=None, timeout=25):
    """'display current-configuration' / 'show running-config' gibi cok
    satirli ciktilar icin. _run_cmd'nin basit '>'/'#' arama mantigi bu
    komutlarda YANLIS sonuç veriyordu: config'in HER BLOGU '#' ile bitiyor
    (Huawei ayraci), bu yuzden komut daha tamamlanmadan ilk blokta 'bitti'
    saniliyordu ve gerisi soket'te akmaya devam edip sonraki komutlarla
    karisiyordu.

    Sadece PROMPT_RE'ye bakmak da yetmiyor: config'in basindaki surum
    bilgisi satiri (orn. '[V200R010C10SPC700]') de ayni formata uyuyor ve
    gercek promptla karisiyor. Bu yuzden expected_hostname verildiginde
    (login sirasinda zaten okunan gercek hostname) SADECE o ismi tasiyan
    satir gercek prompt sayilir; verilmezse (hostname bilinmiyorsa) eski
    genel PROMPT_RE kontrolüne dusulur."""
    tn.drain()
    tn.write(command)
    end_time = time.time() + timeout
    while True:
        lines = tn.text_buffer.splitlines()
        if lines:
            son_satir = lines[-1].strip("\r").strip()
            m = PROMPT_RE.match(son_satir)
            if m and (not expected_hostname or m.group(1) == expected_hostname):
                return tn.text_buffer
        remaining = end_time - time.time()
        if remaining <= 0:
            return tn.text_buffer
        tn.sock.settimeout(min(remaining, 1.0))
        try:
            chunk = tn.sock.recv(4096)
        except socket.timeout:
            continue
        except OSError:
            return tn.text_buffer
        if not chunk:
            return tn.text_buffer
        tn._process(chunk)


KOMUT_HATASI_IFADELERI = [
    "unrecognized command",
    "invalid input",
    "ambiguous command",
    "incomplete command",
]


def _komut_hatasi_mi(output):
    """Cihazin komutu tanimadigini gosteren spesifik hata ifadelerini arar.
    Genel 'error' kelimesi kullanilmaz cunku bazi ciktilarda (orn. Huawei'nin
    'ip interface brief' lejantindaki '(ed): error down' satiri gibi) zararsiz
    sekilde gecebiliyor ve yanlis marka tespitine sebep oluyordu."""
    low = output.lower()
    return any(k in low for k in KOMUT_HATASI_IFADELERI)


def _detect_brand_and_brief(tn, timeout):
    """Huawei/Cisco ayrimini yapar ve 'ip interface brief' ciktisini doner."""
    _run_cmd(tn, "screen-length 0 temporary", timeout=timeout)  # Huawei sayfalama kapama

    out = _run_cmd(tn, "display ip interface brief", timeout=timeout)
    if not _komut_hatasi_mi(out):
        return "huawei", out

    # Huawei komutu calismadi -> Cisco varsayimiyla dene
    _run_cmd(tn, "terminal length 0", timeout=timeout)  # Cisco sayfalama kapama
    out2 = _run_cmd(tn, "show ip interface brief", timeout=timeout)
    return "cisco", out2


def _parse_vlan_ips(brief_output):
    """'... interface X.<vlan> <ip>/<mask> ...' veya 'Vlanif<vlan> <ip>/<mask> ...'
    seklindeki satirlardan {vlan_no: ip} cikarir."""
    result = {}
    for line in brief_output.splitlines():
        m = IFACE_IP_RE.search(line + " ")
        if m:
            vlan_no = int(m.group(1) or m.group(2))
            ip = m.group(3)
            if ip != "0.0.0.0" and vlan_no in VLAN_TANIMLARI:
                result[vlan_no] = ip
    return result


def _parse_vlan_durumlari(brief_output):
    """Ayni 'ip interface brief' ciktisindan VLAN'larin Physical/Protocol
    durumunu UP/DOWN olarak cikarir (yonergedeki Sekil 1-1: Router-Switch
    Up-Link kablo sorunu Physical/Protocol DOWN gorunur). Basit yaklasim:
    IP'yi tasiyan satirda 'down' kelimesi geciyorsa DOWN sayilir."""
    result = {}
    for line in brief_output.splitlines():
        m = IFACE_IP_RE.search(line + " ")
        if m:
            vlan_no = int(m.group(1) or m.group(2))
            if vlan_no in VLAN_TANIMLARI:
                result[vlan_no] = "DOWN" if "down" in line.lower() else "UP"
    return result


# Router'in gercek (WAN/NAT) IP'si genelde LoopBack0/Loopback0 uzerinde
# tanimlidir ve 95.1.x.x formatindadir (bkz. yonerge: "Gercek IP'miz
# (95.1.58.111) ile 4.2.2.2'ye ping atabiliyorsak internete cikisimiz vardir").
LOOPBACK0_RE = re.compile(r"loopback0(?:\.\d+)?\s+(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})", re.IGNORECASE)


def _parse_reel_ip(brief_output):
    """'ip interface brief' ciktisindan LoopBack0/Loopback0 (gercek WAN/NAT
    IP'si) adresini bulur; yoksa None doner."""
    for line in brief_output.splitlines():
        m = LOOPBACK0_RE.search(line)
        if m:
            ip = m.group(1)
            if ip != "0.0.0.0":
                return ip
    return None


ADSL_MODEM_IPLER = ("192.168.1.1", "192.168.1.255")


def _harici_modem_tespiti(arp_output):
    """Yonergeye gore ARP tablosunda 192.168.1.1 veya 192.168.1.255 IP'si
    varsa Fatih agina disaridan bir ADSL modem takilmis demektir.
    NOT: Basit substring arama (`"192.168.1.1" in output`) 192.168.1.173
    gibi IP'lerin ICINDE de "192.168.1.1" gectigi icin yanlis pozitif
    veriyordu - bu yuzden tum IP'ler ayiklanip TAM esitlik ile karsilastirilir."""
    bulunan_ipler = set(IP_RE.findall(arp_output))
    return any(ip in bulunan_ipler for ip in ADSL_MODEM_IPLER)


NAT_CONFIG_RE = re.compile(r"nat outbound|ip nat inside source", re.IGNORECASE)


def _nat_konfigure_mi(config_output):
    """Router'in NAT yapip yapmadigini konfigurasyonun kendisinden anlar
    (Huawei: 'nat outbound ...', Cisco: 'ip nat inside source ...').

    NOT: Onceki surum "display nat session all" / "show ip nat translations"
    ile ANLIK session tablosuna bakiyordu - test aninda trafik yoksa tablo
    bos gorunup NAT calismiyor sanilabiliyordu (yanlis alarm). Konfigurasyona
    bakmak bu sorunu ortadan kaldirir ve zaten baska bir kontrol icin
    cekilmis olan 'display current-configuration' / 'show running-config'
    ciktisini tekrar kullandigi icin ekstra komut calistirmaya da gerek
    kalmaz."""
    return bool(NAT_CONFIG_RE.search(config_output))


# Huawei'de donanim/kart sagligi "display device" komutuyla goruntulenir.
# Buyuk (chassis tabanli) modellerde SRU (Switching & Routing Unit - ana
# islemci karti) burada ayri bir slot olarak listelenir; kucuk box tipi
# router'larda (orn. AR1220F) tek bir ana unite vardir ama ayni "Alarm"
# sutunu genel donanim sagligini gosterir. Bu yuzden SRU'ya ozel arama
# yerine, herhangi bir kartin Alarm durumu "Normal" degilse donanim
# hatasi/alarm var sayilir - bu SRU'yu da (varsa) kapsar.
DONANIM_ALARM_IFADELERI = ("major", "minor", "critical", "fault", "abnormal")


def _donanim_komutu(brand):
    return "display device" if brand == "huawei" else None


def _donanim_alarm_var_mi(device_output):
    """None: kontrol yapilamadi (orn. Cisco'da komut yok/farkli).
    True/False: donanim alarmi var/yok."""
    if device_output is None:
        return None
    if _komut_hatasi_mi(device_output):
        return None
    low = device_output.lower()
    return any(k in low for k in DONANIM_ALARM_IFADELERI)


ALARM_SEVIYE_ONCELIK = {"critical": 4, "major": 3, "minor": 2, "warning": 1}


def _alarm_active_analiz(alarm_output):
    """'display alarm active' ciktisindaki aktif alarmlarin EN YUKSEK
    seviyesini bulur. Basit "cikti bos mu degil mi" kontrolu yaniltici
    oluyordu: orn. bir portun optik sinyali dusukse (hwOpticalInvalid,
    Warning seviyesi) bu router'in ARIZALI oldugu anlamina gelmez - kart
    seviyesinde saglikli olabilir, sadece kullanilmayan/zayif bir port
    uyarisi olabilir. Bu yuzden severity ayirt edilir.

    Donus: (en_yuksek_seviye, alarm_sayisi). en_yuksek_seviye None ise
    kontrol yapilamadi (komut desteklenmiyor); "" ise alarm yok; aksi
    halde "Critical"/"Major"/"Minor"/"Warning"."""
    if alarm_output is None or _komut_hatasi_mi(alarm_output):
        return None, 0
    seviyeler = []
    for line in alarm_output.splitlines():
        low = line.lower()
        for sev in ("critical", "major", "minor", "warning"):
            if f"/{sev}/" in low:
                seviyeler.append(sev)
                break
    if not seviyeler:
        return "", 0
    en_yuksek = max(seviyeler, key=lambda s: ALARM_SEVIYE_ONCELIK.get(s, 0))
    return en_yuksek.capitalize(), len(seviyeler)


def _ping_cmd(brand, source_ip, dest_ip):
    # Varsayilan 5 paket yerine 2 paket gonderilir - script'i ciddi
    # yavaslatan asil kalem buydu (satir basina 9 ping var, paket kaybi
    # olan VLAN'larda her biri ~10sn surebiliyordu). 2 paket VAR/YOK
    # ayrimi icin yeterlidir (0% ya da 100% kayip degismez), sadece KISMI
    # sonuclarin yuzdesi daha kaba olur (orn. %50 kayip = 1/2 paket).
    if brand == "huawei":
        return f"ping -c 2 -a {source_ip} {dest_ip}"
    return f"ping {dest_ip} source {source_ip} repeat 2"


CISCO_SUCCESS_RATE_RE = re.compile(r"success rate is (\d+) percent", re.IGNORECASE)


def _ping_sonucu(output):
    # Huawei formati: "0.00% packet loss" / "100.00% packet loss"
    m = PACKET_LOSS_RE.search(output)
    if m:
        loss = float(m.group(1))
        if loss == 0:
            return "VAR"
        if loss >= 100:
            return "YOK"
        return f"KISMI(%{loss:g} kayip)"

    # Cisco formati: "Success rate is 100 percent (5/5), ..."
    m2 = CISCO_SUCCESS_RATE_RE.search(output)
    if m2:
        rate = int(m2.group(1))
        if rate == 100:
            return "VAR"
        if rate == 0:
            return "YOK"
        return f"KISMI(%{rate} basarili)"

    low = output.lower()
    if "reply from" in low or "bytes from" in low:
        return "VAR"
    # NOT: Cisco ping ciktisinda "timeout is 2 seconds" gibi ping'in kendi
    # ayarini belirten zararsiz bir "timeout" ifadesi de gecebiliyor, bu
    # yuzden sadece "istek zaman asimina ugradi" anlamina gelen NET
    # ifadeler aranir (genel "timeout" kelimesi ARANMAZ).
    if "request time out" in low or "request timed out" in low or "unreachable" in low:
        return "YOK"
    return "BELIRSIZ"


def _arp_command(brand):
    return "display arp dynamic" if brand == "huawei" else "show arp"


def _arp_cihaz_sayisi(output):
    """ARP tablosundaki (IP icerern) satirlari sayar. Sadece komut yankisi
    satiri elenir; onceki surumde 'arp' kelimesi geciyor diye satir atlanan
    bir filtre vardi ama Cisco ciktisindaki 'ARPA' tipi de bu kelimeyi
    icerdigi icin TUM satirlari yanlislikla eliyordu - bu yuzden kaldirildi."""
    count = 0
    for line in output.splitlines():
        low = line.strip().lower()
        if low.startswith("display arp") or low.startswith("show arp"):
            continue  # komut yankisi
        if IP_RE.search(line):
            count += 1
    return count


def _config_komutu(brand):
    return "display current-configuration" if brand == "huawei" else "show running-config"


# DHCP pool bloklarini (Huawei "ip pool <isim>" / Cisco "ip dhcp pool <isim>")
# ayirip, icindeki gateway IP'sini ve option 43/148 varligini bulur.
# Bu option'lar AP'lere WLC adresini soyler; olmazsa kablosuz (AP) calismaz.
# NOT: Eski router'larda sadece option 43 kullanilir. Yeni formattaki
# router'larda (VLAN1'i olanlarda) option 148 de kullanilabiliyor - bu
# yuzden yeni router'larda hem VLAN1 hem VLAN10 icin 43 VEYA 148 aranir.
POOL_BLOCK_RE = re.compile(
    r"ip (?:dhcp )?pool\s+\S+\r?\n(.*?)(?=\r?\n[#!]|\Z)", re.DOTALL | re.IGNORECASE
)
GATEWAY_RE = re.compile(r"(?:gateway-list|default-router)\s+(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})")
DHCP_OPTION_ESKI_RE = re.compile(r"option\s+43\b", re.IGNORECASE)
DHCP_OPTION_YENI_RE = re.compile(r"option\s+(43|148)\b", re.IGNORECASE)


def _dhcp_option_kontrolu(config_output, vlan_ips):
    """VLAN1 ve VLAN10'un DHCP pool'unda option 43 (eski router) veya
    option 43/148'den biri (yeni router) var mi kontrol eder (AP'lerin
    WLC'yi bulmasi icin gerekli). Sonuc: {vlan_no: "VAR"/"YOK"/"IP BULUNAMADI"}.

    Her VLAN kendi IP'sinin bulunup bulunmadigina gore BAGIMSIZ kontrol
    edilir: eski router'larda VLAN1 yoktur, IP bulunamayacagi icin otomatik
    "IP BULUNAMADI" (Excel'de "-") yazilir, ama VLAN10 her router'da vardir
    ve HER ZAMAN kontrol edilir. Yeni router'larda (VLAN1'i de olanlarda)
    hem VLAN1 hem VLAN10 kontrol edilir, ikisinde de 43/148 aranir."""
    sonuc = {}
    pool_bloklari = POOL_BLOCK_RE.findall(config_output)
    yeni_router_mu = bool(vlan_ips.get(1))
    option_regex = DHCP_OPTION_YENI_RE if yeni_router_mu else DHCP_OPTION_ESKI_RE
    for vlan_no in (1, 10):
        gw_ip = vlan_ips.get(vlan_no)
        if not gw_ip:
            sonuc[vlan_no] = "IP BULUNAMADI"
            continue
        bulundu = False
        for pool_body in pool_bloklari:
            gw_match = GATEWAY_RE.search(pool_body)
            if gw_match and gw_match.group(1) == gw_ip:
                if option_regex.search(pool_body):
                    bulundu = True
                break
        sonuc[vlan_no] = "VAR" if bulundu else "YOK"
    return sonuc


def kontrol_et(ip, username, password, port=23, timeout=10):
    """
    Tum kontrol adimlarini uygular. Sonuc: dict.
    """
    sonuc = {
        "erisim": None,
        "okul_adi": "",
        "marka": "",
        "vlan_sonuclari": {},  # vlan_no -> "VAR"/"YOK"/"IP YOK"/...
        "arp_sayisi": None,
        "wlc_erisim": "",
        "eba_erisim": "",
        "dhcp_option_sonuclari": {},  # {1: "VAR"/"YOK"/"IP BULUNAMADI", 10: ...}
        "interface_durumlari": {},  # {vlan_no: "UP"/"DOWN"}
        "harici_modem": False,
        "nat_calisiyor": None,
        "donanim_alarm": None,  # True/False/None (SRU/kart alarmi, sadece Huawei)
        "alarm_seviye": None,  # ""/"Warning"/"Minor"/"Major"/"Critical"/None
        "alarm_sayisi": 0,
        "reel_ip": None,  # LoopBack0 (WAN/NAT) IP'si, sadece tum VLAN'lar kapaliysa doldurulur
        "reel_ip_erisim": "",  # VAR/YOK/BELIRSIZ/REEL IP BULUNAMADI, ayni sekilde
        "not": "",
    }
    tn = None
    try:
        tn = MiniTelnet(ip, port, timeout=timeout)

        ok, detay = _login(tn, username, password, timeout)
        if not ok:
            sonuc["erisim"] = f"ERISIM YOK: {detay}"
            return sonuc
        sonuc["erisim"] = "ERISIM VAR"
        sonuc["okul_adi"] = _find_prompt(detay) or ""

        brand, brief = _detect_brand_and_brief(tn, timeout)
        sonuc["marka"] = brand.upper()

        vlan_ips = _parse_vlan_ips(brief)
        if not vlan_ips:
            sonuc["not"] = "VLAN interface IP'leri 'ip interface brief' ciktisindan okunamadi"
        sonuc["interface_durumlari"] = _parse_vlan_durumlari(brief)

        # DHCP pool konfigurasyonunda option 43 var mi (kablosuz/AP
        # icin gerekli). "display current-configuration" buyuk bir cikti
        # oldugu icin genis bir timeout verilir.
        config_out = _run_config_cmd(
            tn, _config_komutu(brand), expected_hostname=sonuc["okul_adi"] or None, timeout=max(timeout, 25)
        )
        sonuc["dhcp_option_sonuclari"] = _dhcp_option_kontrolu(config_out, vlan_ips)

        # 2 paketlik ping (bkz. _ping_cmd) genelde ~4-6 saniyede biter;
        # kisa timeout ping ciktisini yarim keserse o kalinti bir sonraki
        # komutun yanitiyla karisip yanlis sonuca (orn. ARP sayisi 0) sebep
        # oluyordu - bu yuzden yine de makul bir pay birakilir.
        ping_timeout = max(timeout, 8)

        for vlan_no in VLAN_TANIMLARI:
            src_ip = vlan_ips.get(vlan_no)
            if not src_ip:
                sonuc["vlan_sonuclari"][vlan_no] = "IP BULUNAMADI"
                continue
            out = _run_cmd(tn, _ping_cmd(brand, src_ip, PING_HEDEFI), timeout=ping_timeout)
            sonuc["vlan_sonuclari"][vlan_no] = _ping_sonucu(out)

        # Hicbir VLAN'in interneti yoksa, router'in gercek (WAN/NAT) IP'si
        # uzerinden genel bir internet cikisi testi yapilir (yonergedeki
        # "gercek IP ile 4.2.2.2'ye ping" kontrolu). Bu, sorunun router/MPLS
        # seviyesinde mi (reel IP de pingsizse) yoksa switch/VLAN seviyesinde
        # mi (reel IP pinglenebiliyorsa) oldugunu ayirt etmeye yardimci olur.
        hicbir_vlan_calismıyor = not any(v == "VAR" for v in sonuc["vlan_sonuclari"].values())
        if hicbir_vlan_calismıyor:
            reel_ip = _parse_reel_ip(brief)
            sonuc["reel_ip"] = reel_ip
            if reel_ip:
                reel_out = _run_cmd(tn, _ping_cmd(brand, reel_ip, PING_HEDEFI), timeout=ping_timeout)
                sonuc["reel_ip_erisim"] = _ping_sonucu(reel_out)
            else:
                sonuc["reel_ip_erisim"] = "REEL IP BULUNAMADI"

        arp_out = _run_cmd(tn, _arp_command(brand), timeout=timeout)
        sonuc["arp_sayisi"] = _arp_cihaz_sayisi(arp_out)
        sonuc["harici_modem"] = _harici_modem_tespiti(arp_out)

        # NAT kontrolu: konfigurasyonun kendisinde NAT tanimli mi (config_out
        # zaten DHCP option kontrolu icin cekilmisti, ekstra komut gerekmez).
        sonuc["nat_calisiyor"] = _nat_konfigure_mi(config_out)

        # Donanim/log alarm kontrolu - sadece Huawei'de yapilir. Ikisi
        # birlestirilir: "display device" (kart/SRU durumu - Normal degilse
        # ciddi bir kart arizasidir) ve "display alarm active" (BPDU dahil
        # her turlu aktif sistem alarmi, severity'ye gore degerlendirilir -
        # Warning/Minor seviyesi "router bozuk" degil, "hafif uyari" sayilir).
        donanim_cmd = _donanim_komutu(brand)
        if donanim_cmd:
            donanim_out = _run_config_cmd(
                tn, donanim_cmd, expected_hostname=sonuc["okul_adi"] or None, timeout=max(timeout, 15)
            )
            kart_alarm = _donanim_alarm_var_mi(donanim_out)

            alarm_out = _run_config_cmd(
                tn, "display alarm active", expected_hostname=sonuc["okul_adi"] or None, timeout=max(timeout, 15)
            )
            alarm_seviye, alarm_sayisi = _alarm_active_analiz(alarm_out)
            sonuc["alarm_seviye"] = alarm_seviye
            sonuc["alarm_sayisi"] = alarm_sayisi

            # Sadece Major/Critical seviyesi "ciddi ariza" sayilir; kart
            # alarmi (display device) zaten her zaman ciddi kabul edilir.
            ciddi_sistem_alarm = alarm_seviye in ("Major", "Critical")

            if kart_alarm is True or ciddi_sistem_alarm:
                sonuc["donanim_alarm"] = True
            elif kart_alarm is False:
                sonuc["donanim_alarm"] = False
            else:
                sonuc["donanim_alarm"] = None
        else:
            sonuc["donanim_alarm"] = None

        # WLC kontrolu: "AP'ler WLC'ye gidiyor mu?" -> VLAN10 (Yonetim/AP) IP'sinden
        yonetim_ip = vlan_ips.get(10)
        if yonetim_ip:
            wlc_out = _run_cmd(tn, _ping_cmd(brand, yonetim_ip, WLC_IP), timeout=ping_timeout)
            sonuc["wlc_erisim"] = _ping_sonucu(wlc_out)
        else:
            sonuc["wlc_erisim"] = "VLAN10 IP BULUNAMADI"

        # EBA kontrolu: "ET'ler EBA sunucularina gidiyor mu?" -> VLAN11 (Tahta) IP'sinden
        tahta_ip = vlan_ips.get(11)
        if tahta_ip:
            eba_out = _run_cmd(tn, _ping_cmd(brand, tahta_ip, EBA_IP), timeout=ping_timeout)
            sonuc["eba_erisim"] = _ping_sonucu(eba_out)
        else:
            sonuc["eba_erisim"] = "VLAN11 IP BULUNAMADI"

        return sonuc

    except socket.timeout:
        sonuc["erisim"] = "ERISIM YOK: Zaman asimi (cihaza ulasilamadi)"
        return sonuc
    except ConnectionRefusedError:
        sonuc["erisim"] = "ERISIM YOK: Baglanti reddedildi (telnet portu kapali olabilir)"
        return sonuc
    except OSError as e:
        sonuc["erisim"] = f"ERISIM YOK: Baglanti hatasi: {e}"
        return sonuc
    except Exception as e:  # noqa: BLE001
        sonuc["erisim"] = f"HATA: {e}"
        return sonuc
    finally:
        if tn:
            tn.close()


def ozet_cikar(sonuc):
    """Tum kontrol sonuclarina bakip tek satirlik bir genel durum ozeti
    uretir: sorun yoksa 'Sorun yok', varsa hangi sorunlar oldugunu yazar."""
    if not sonuc["erisim"] or not sonuc["erisim"].startswith("ERISIM VAR"):
        return sonuc["erisim"] or "Bilinmeyen hata"

    sorunlar = []

    for vlan_no, sonuc_metni in sonuc["vlan_sonuclari"].items():
        kisa_isim = VLAN_KISA_ISIM.get(vlan_no, f"VLAN{vlan_no}")
        if sonuc_metni == "IP BULUNAMADI":
            continue  # bu VLAN cihazda tanimli degil, sorun sayilmaz
        if sonuc_metni != "VAR":
            sorunlar.append(f"{kisa_isim} interneti yok ({sonuc_metni})")

    if sonuc.get("reel_ip_erisim"):
        if sonuc["reel_ip_erisim"] == "VAR":
            sorunlar.append(
                f"Hicbir VLAN'in interneti yok AMA router'in gercek IP'si ({sonuc.get('reel_ip')}) "
                "internete cikabiliyor - sorun MPLS/router degil, switch/VLAN/DHCP seviyesinde olabilir"
            )
        elif sonuc["reel_ip_erisim"] == "REEL IP BULUNAMADI":
            sorunlar.append("Hicbir VLAN'in interneti yok, router'in gercek (LoopBack0) IP'si de bulunamadi")
        else:
            sorunlar.append(
                f"Hicbir VLAN'in interneti yok, router'in gercek IP'si ({sonuc.get('reel_ip')}) ile de "
                f"internete cikilamiyor ({sonuc['reel_ip_erisim']}) - MPLS/Turk Telekom kaynakli olabilir"
            )

    if sonuc["arp_sayisi"] == 0:
        sorunlar.append("ARP tablosunda hic cihaz yok (switch/AP bagli olmayabilir)")

    if sonuc["wlc_erisim"] and sonuc["wlc_erisim"] not in ("VAR", "VLAN10 IP BULUNAMADI"):
        sorunlar.append(f"WLC erisimi yok ({sonuc['wlc_erisim']})")

    if sonuc["eba_erisim"] and sonuc["eba_erisim"] not in ("VAR", "VLAN11 IP BULUNAMADI"):
        sorunlar.append(f"EBA erisimi yok ({sonuc['eba_erisim']})")

    dhcp_kisa_isim = {1: "VLAN1", 10: "VLAN10"}
    for vlan_no, deger in sonuc.get("dhcp_option_sonuclari", {}).items():
        if deger == "IP BULUNAMADI":
            continue  # bu VLAN cihazda tanimli degil, sorun sayilmaz
        if deger != "VAR":
            sorunlar.append(
                f"{dhcp_kisa_isim.get(vlan_no, f'VLAN{vlan_no}')} DHCP pool'unda option 43 yok (kablosuz/AP calismayabilir)"
            )

    for vlan_no, durum in sonuc.get("interface_durumlari", {}).items():
        if durum == "DOWN":
            kisa_isim = VLAN_KISA_ISIM.get(vlan_no, f"VLAN{vlan_no}")
            sorunlar.append(f"{kisa_isim} interface DOWN (Router-Switch Up-Link kablo sorunu olabilir)")

    if sonuc.get("harici_modem"):
        sorunlar.append("ARP'ta harici ADSL modem/router tespit edildi (192.168.1.x)")

    if sonuc.get("nat_calisiyor") is False:
        sorunlar.append("Router NAT yapmiyor gibi gorunuyor (NAT session/translation bulunamadi)")

    # Sadece Major/Critical seviyesi (kart arizasi dahil) sorun sayilir.
    # Warning/Minor seviyeler bilerek RAPORLANMAZ - kafa karistirdigi icin
    # (orn. kullanilmayan bir portta dusuk optik sinyal gibi onemsiz
    # durumlar router'in bozuk oldugu izlenimi veriyordu).
    if sonuc.get("donanim_alarm") is True:
        seviye = sonuc.get("alarm_seviye") or ""
        sorunlar.append(
            f"Router donanim/kart alarmi tespit edildi (SRU dahil, seviye:{seviye or 'Kart Arizasi'} - "
            f"'display device'/'display alarm active' ciktisina bakin, router ariza yapmis olabilir)"
        )

    if sonuc["not"]:
        sorunlar.append(sonuc["not"])

    if not sorunlar:
        return "Sorun yok"
    return "; ".join(sorunlar)


TEKRAR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # acik kirmizi
TEKRAR_FONT = Font(color="9C0006")  # koyu kirmizi yazi


def tekrarlari_isaretle_ve_tasi(ws, col):
    """Ayni IP birden fazla satirda geciyorsa: ilk gorulen satir oldugu gibi
    kalir, sonraki tekrarlar kirmiziya boyanip 'TEKRAR EDEN IP' notu ile
    isaretlenir ve listenin en altina tasinir (boylece kontrol listesi
    temiz kalir, tekrar eden satirlar goze batar ve ayrica test edilmez)."""
    ip_col = col.get("IP Adresi")
    if not ip_col:
        return

    gorulen_ip = set()
    tekrar_eden_satirlar = []  # buyukten kucuge, silmeden once toplanir
    for row in range(2, ws.max_row + 1):
        ip_val = ws.cell(row=row, column=ip_col).value
        if not ip_val:
            continue
        ip = str(ip_val).strip()
        if not ip:
            continue
        if ip in gorulen_ip:
            tekrar_eden_satirlar.append(row)
        else:
            gorulen_ip.add(ip)

    if not tekrar_eden_satirlar:
        return

    # Once verileri oku, sonra buyuk satir numarasindan kucuge dogru sil
    # (kucukten silinirse sonraki satir numaralari kayar ve yanlis satir silinir).
    tasinacaklar = []
    for row in sorted(tekrar_eden_satirlar, reverse=True):
        degerler = [ws.cell(row=row, column=c).value for c in range(1, ws.max_column + 1)]
        tasinacaklar.append(degerler)
        ws.delete_rows(row, 1)

    sorun_col = col.get("Sorun / Genel Durum")
    erisim_col = col.get("Erisim Durumu")
    for degerler in reversed(tasinacaklar):  # orijinal sirayla en alta ekle
        yeni_row = ws.max_row + 1
        for c_idx, deger in enumerate(degerler, start=1):
            cell = ws.cell(row=yeni_row, column=c_idx)
            cell.value = deger
            cell.fill = TEKRAR_FILL
            cell.font = TEKRAR_FONT
        if sorun_col:
            ws.cell(row=yeni_row, column=sorun_col).value = "TEKRAR EDEN IP (listede baska bir satirda da var)"
        if erisim_col:
            ws.cell(row=yeni_row, column=erisim_col).value = "ATLANDI: Tekrar eden IP"

    print(f"UYARI: {len(tasinacaklar)} tekrar eden IP satiri kirmiziya boyanip listenin en altina tasindi.")


# Network I/O bound bir is oldugu icin (CPU degil) coklu thread GIL'e
# ragmen gercek bir hizlanma saglar - ayni anda birden fazla router'a
# baglanip bekleme sureleri ortusur. 8 es zamanli baglanti guvenli bir
# sayidir (tek makineden farkli IP'lere gidiyor, hicbir router'in VTY
# limitini zorlamaz).
MAX_WORKERS = 8

_BOS_SONUC_SABLONU = {
    "erisim": None, "okul_adi": "", "marka": "", "vlan_sonuclari": {},
    "arp_sayisi": None, "wlc_erisim": "", "eba_erisim": "",
    "dhcp_option_sonuclari": {}, "interface_durumlari": {},
    "harici_modem": False, "nat_calisiyor": None, "donanim_alarm": None,
    "alarm_seviye": None, "alarm_sayisi": 0, "reel_ip": None,
    "reel_ip_erisim": "", "not": "",
}


def _sonucu_excel_yaz(ws, col, row, sonuc):
    """Bir kontrol_et() sonucunu Excel satirina yazar ve loglar."""
    okul_cell = ws.cell(row=row, column=col["Okul / Aciklama"]) if "Okul / Aciklama" in col else None
    if okul_cell is not None and not (okul_cell.value and str(okul_cell.value).strip()) and sonuc["okul_adi"]:
        okul_cell.value = sonuc["okul_adi"]

    ws.cell(row=row, column=col["Erisim Durumu"]).value = sonuc["erisim"]
    if "Marka" in col:
        ws.cell(row=row, column=col["Marka"]).value = sonuc["marka"]
    for vlan_no, sutun_adi in VLAN_TANIMLARI.items():
        if sutun_adi in col:
            deger = sonuc["vlan_sonuclari"].get(vlan_no, "")
            # "IP BULUNAMADI" o VLAN'in bu cihazda tanimli olmadigi
            # anlamina gelir (orn. VLAN1 eski routerlarda yok) - sorun
            # degildir, Excel'de sade bir "-" ile gosterilir.
            ws.cell(row=row, column=col[sutun_adi]).value = "-" if deger == "IP BULUNAMADI" else deger
    if "ARP Cihaz Sayisi" in col:
        ws.cell(row=row, column=col["ARP Cihaz Sayisi"]).value = sonuc["arp_sayisi"]
    if "WLC Erisimi" in col:
        ws.cell(row=row, column=col["WLC Erisimi"]).value = sonuc["wlc_erisim"]
    if "EBA Erisimi" in col:
        ws.cell(row=row, column=col["EBA Erisimi"]).value = sonuc["eba_erisim"]
    dhcp_sutunlari = {1: "VLAN1 Option43 (Kablosuz)", 10: "VLAN10 Option43 (Kablosuz)"}
    for vlan_no, sutun_adi in dhcp_sutunlari.items():
        if sutun_adi in col:
            deger = sonuc["dhcp_option_sonuclari"].get(vlan_no, "")
            ws.cell(row=row, column=col[sutun_adi]).value = "-" if deger == "IP BULUNAMADI" else deger
    if "Reel IP Internet Cikisi" in col:
        reel_erisim = sonuc.get("reel_ip_erisim")
        reel_ip = sonuc.get("reel_ip")
        if reel_erisim:
            deger_metni = f"{reel_erisim} ({reel_ip})" if reel_ip else reel_erisim
        else:
            deger_metni = "-"  # test edilmedi (en az bir VLAN calisiyordu)
        ws.cell(row=row, column=col["Reel IP Internet Cikisi"]).value = deger_metni
    if "Harici ADSL Modem" in col:
        ws.cell(row=row, column=col["Harici ADSL Modem"]).value = "VAR" if sonuc.get("harici_modem") else "YOK"
    if "NAT Calisiyor mu" in col:
        ws.cell(row=row, column=col["NAT Calisiyor mu"]).value = "VAR" if sonuc.get("nat_calisiyor") else "YOK"
    if "Donanim/Log Alarmi (SRU-BPDU vb.)" in col:
        donanim_deger = sonuc.get("donanim_alarm")
        # Sadece Major/Critical (veya kart arizasi) gosterilir; Warning/
        # Minor seviyeler bilerek "YOK" gibi gosterilir - kafa karistirmasin.
        if donanim_deger is True:
            seviye = sonuc.get("alarm_seviye") or "Kart Arizasi"
            deger_metni = f"VAR - {seviye} (ARIZALI OLABILIR)"
        elif donanim_deger is False:
            deger_metni = "YOK"
        else:
            deger_metni = "N/A (Cisco'da kontrol edilmiyor)"
        ws.cell(row=row, column=col["Donanim/Log Alarmi (SRU-BPDU vb.)"]).value = deger_metni
    if "Interface Durumu (DOWN olanlar)" in col:
        down_vlanlar = [
            VLAN_KISA_ISIM.get(v, f"VLAN{v}")
            for v, d in sonuc.get("interface_durumlari", {}).items()
            if d == "DOWN"
        ]
        ws.cell(row=row, column=col["Interface Durumu (DOWN olanlar)"]).value = (
            ", ".join(down_vlanlar) if down_vlanlar else "-"
        )
    if "Not" in col:
        ws.cell(row=row, column=col["Not"]).value = sonuc["not"]

    ozet = ozet_cikar(sonuc)
    if "Sorun / Genel Durum" in col:
        ws.cell(row=row, column=col["Sorun / Genel Durum"]).value = ozet

    print(f"    -> OZET: {ozet}")
    print(f"       Erisim:{sonuc['erisim']}  Okul:{sonuc['okul_adi']}  Marka:{sonuc['marka']}  VLAN:{sonuc['vlan_sonuclari']}  ARP:{sonuc['arp_sayisi']}  WLC:{sonuc['wlc_erisim']}  EBA:{sonuc['eba_erisim']}  DHCP-Opt:{sonuc['dhcp_option_sonuclari']}  DonanimAlarm:{sonuc.get('donanim_alarm')}  AlarmSeviye:{sonuc.get('alarm_seviye')}({sonuc.get('alarm_sayisi')})  ReelIP:{sonuc.get('reel_ip')}={sonuc.get('reel_ip_erisim')}")


def main():
    retry_failed = "--retry-failed" in sys.argv
    numeric_args = [a for a in sys.argv[1:] if a != "--retry-failed"]
    limit = int(numeric_args[0]) if numeric_args else None

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    col = {name: idx + 1 for idx, name in enumerate(headers) if name}

    required = ["IP Adresi", "Kullanici Adi", "Sifre", "Erisim Durumu"]
    for r in required:
        if r not in col:
            raise SystemExit(f"Excel'de '{r}' sutunu bulunamadi. Sablonu bozmadigindan emin ol.")

    tekrarlari_isaretle_ve_tasi(ws, col)
    wb.save(EXCEL_PATH)

    # Once islenecek satirlari topla (Excel okuma islemi tek thread'de,
    # network I/O baslamadan once yapilir).
    gorevler = []  # (row, ip, username, password, port)
    for row in range(2, ws.max_row + 1):
        if limit is not None and len(gorevler) >= limit:
            break

        ip_cell = ws.cell(row=row, column=col["IP Adresi"]).value
        if not ip_cell or not str(ip_cell).strip():
            continue

        existing = ws.cell(row=row, column=col["Erisim Durumu"]).value
        if existing and str(existing).startswith("ATLANDI"):
            continue  # tekrar eden IP olarak isaretlenmis, hic islenmez
        if existing and str(existing).startswith("ERISIM VAR"):
            continue  # zaten basariyla erisilmis, atla
        if existing and not retry_failed:
            continue  # daha once denenmis, --retry-failed verilmemis, atla

        ip = str(ip_cell).strip()
        # Excel'de kullanici adi/sifre bos birakilmissa varsayilanlar kullanilir.
        username = str(ws.cell(row=row, column=col["Kullanici Adi"]).value or "").strip() or VARSAYILAN_KULLANICI_ADI
        password = str(ws.cell(row=row, column=col["Sifre"]).value or "").strip() or VARSAYILAN_SIFRE
        port_val = ws.cell(row=row, column=col.get("Port (bos=23)", 0)).value if col.get("Port (bos=23)") else None
        port = int(port_val) if port_val else 23
        gorevler.append((row, ip, username, password, port))

    if not gorevler:
        print("Islenecek yeni satir yok (hepsi zaten basarili/atlanmis).")
        print("\nTamamlandi. Dosya guncellendi:", EXCEL_PATH)
        return

    print(f"{len(gorevler)} IP, {MAX_WORKERS} es zamanli baglanti ile kontrol edilecek...\n")

    # Network I/O (Telnet) is'i worker thread'lerde paralel yapilir; Excel'e
    # yazma (ws.cell/wb.save) SADECE ana thread'de yapilir - openpyxl
    # thread-safe degildir, bu yuzden worker'lar sadece kontrol_et() sonucu
    # doner, hicbir worker dogrudan Excel'e dokunmaz.
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_to_gorev = {
            executor.submit(kontrol_et, ip, username, password, port=port): (row, ip)
            for row, ip, username, password, port in gorevler
        }
        tamamlanan = 0
        for future in as_completed(future_to_gorev):
            row, ip = future_to_gorev[future]
            tamamlanan += 1
            try:
                sonuc = future.result()
            except Exception as e:  # noqa: BLE001
                sonuc = dict(_BOS_SONUC_SABLONU)
                sonuc["erisim"] = f"HATA: {e}"

            print(f"[Satir {row}] {ip}  ({tamamlanan}/{len(gorevler)} tamamlandi)")
            _sonucu_excel_yaz(ws, col, row, sonuc)
            wb.save(EXCEL_PATH)  # her sonuctan sonra kaydet ki kesinti olursa veri kaybolmasin

    print("\nTamamlandi. Dosya guncellendi:", EXCEL_PATH)


if __name__ == "__main__":
    main()
